import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

class KudirGenerator:
    """Генератор Книги учета доходов и расходов"""
    
    def __init__(self, income_operations):
        self.income_operations = sorted(income_operations, key=lambda x: x['date'])
        self.kudir_data = []
        
    def generate(self):
        """Формирует КУДиР"""
        self._build_kudir()
        return self.kudir_data
    
    def _build_kudir(self):
        """Построение таблицы КУДиР"""
        for idx, op in enumerate(self.income_operations, 1):
            self.kudir_data.append({
                '№ п/п': idx,
                'Дата и номер документа': f"{op['date'].strftime('%d.%m.%Y')} {op['document']}",
                'Содержание операции': op['purpose'][:200],
                'Сумма дохода': op['amount']
            })
    
    def get_quarterly_totals(self):
        """Получение сумм по кварталам"""
        totals = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
        
        for op in self.income_operations:
            quarter = (op['date'].month - 1) // 3 + 1
            totals[quarter] += op['amount']
        
        return totals
    
    def export_to_excel(self, filepath):
        """Экспорт КУДиР в Excel"""
        df = pd.DataFrame(self.kudir_data)
        
        # Форматируем суммы
        df['Сумма дохода'] = df['Сумма дохода'].apply(lambda x: f"{x:,.2f}".replace(",", " "))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "КУДиР"
        
        # Заголовок
        ws['A1'] = "Книга учета доходов и расходов ИП на УСН"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "за 2025 год"
        ws.merge_cells('A2:D2')
        
        # Заголовки колонок
        headers = ['№ п/п', 'Дата и номер документа', 'Содержание операции', 'Сумма дохода']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Данные
        for row_idx, row_data in enumerate(self.kudir_data, 5):
            ws.cell(row=row_idx, column=1, value=row_data['№ п/п'])
            ws.cell(row=row_idx, column=2, value=row_data['Дата и номер документа'])
            ws.cell(row=row_idx, column=3, value=row_data['Содержание операции'])
            ws.cell(row=row_idx, column=4, value=row_data['Сумма дохода'])
        
        # Итог
        total_row = len(self.kudir_data) + 5
        ws.cell(row=total_row, column=3, value="ИТОГО:")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        total_amount = sum(op['amount'] for op in self.income_operations)
        ws.cell(row=total_row, column=4, value=f"{total_amount:,.2f}".replace(",", " "))
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        
        wb.save(filepath)
        
        return total_amount