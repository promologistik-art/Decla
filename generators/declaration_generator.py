import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

class DeclarationGenerator:
    """Генератор декларации по УСН"""
    
    def __init__(self, income_operations, ens_data):
        self.income_ops = income_operations
        self.ens_data = ens_data
        
        # Инициализируем суммы
        self.quarterly_income = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
        self.total_income = 0.0
        
        # Налог
        self.tax_rate = 6
        self.tax_amount = 0.0
        self.tax_payable = 0.0
        
    def calculate(self):
        """Расчет налога"""
        # Считаем доходы по кварталам
        for op in self.income_ops:
            quarter = (op['date'].month - 1) // 3 + 1
            self.quarterly_income[quarter] += op['amount']
            self.total_income += op['amount']
        
        # Налог 6%
        self.tax_amount = self.total_income * self.tax_rate / 100
        
        # Вычет по взносам (только уплаченные в 2025 году)
        # По данным ЕНС: взносы уплачены в 2026 → вычет = 0
        deductible_insurance = self.ens_data.get('insurance_paid', 0)
        
        # Проверяем даты уплаты (только 2025 год)
        paid_in_2025 = 0
        for date in self.ens_data.get('insurance_paid_dates', []):
            if date and date.year == 2025:
                paid_in_2025 += deductible_insurance
        
        # Если взносы не уплачены в 2025, вычет = 0
        if paid_in_2025 == 0:
            self.tax_payable = self.tax_amount
        else:
            self.tax_payable = max(0, self.tax_amount - paid_in_2025)
        
        return {
            'quarterly_income': self.quarterly_income,
            'total_income': self.total_income,
            'tax_amount': self.tax_amount,
            'tax_payable': self.tax_payable
        }
    
    def generate_excel(self, filepath):
        """Генерация декларации в Excel (для печати и проверки)"""
        self.calculate()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Декларация УСН"
        
        # Заголовок
        ws['A1'] = "Налоговая декларация по налогу, уплачиваемому в связи с применением упрощенной системы налогообложения"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "за 2025 год"
        ws.merge_cells('A2:F2')
        
        # Раздел 2.1.1 (доходы)
        ws['A4'] = "Раздел 2.1.1. Доходы"
        ws['A4'].font = Font(bold=True)
        
        ws['A6'] = "Показатель"
        ws['B6'] = "Код строки"
        ws['C6'] = "Значение"
        
        rows_income = [
            ("Доход за 1 квартал", "110", self.quarterly_income[1]),
            ("Доход за полугодие", "111", self.quarterly_income[1] + self.quarterly_income[2]),
            ("Доход за 9 месяцев", "112", self.quarterly_income[1] + self.quarterly_income[2] + self.quarterly_income[3]),
            ("Доход за год", "113", self.total_income),
            ("Налоговая ставка (%)", "120", self.tax_rate),
            ("Сумма налога за 1 квартал", "130", self.quarterly_income[1] * self.tax_rate / 100),
            ("Сумма налога за полугодие", "131", (self.quarterly_income[1] + self.quarterly_income[2]) * self.tax_rate / 100),
            ("Сумма налога за 9 месяцев", "132", (self.quarterly_income[1] + self.quarterly_income[2] + self.quarterly_income[3]) * self.tax_rate / 100),
            ("Сумма налога за год", "133", self.tax_amount),
        ]
        
        for idx, (name, code, value) in enumerate(rows_income, 7):
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=code)
            ws.cell(row=idx, column=3, value=round(value, 2))
        
        # Раздел 1.1 (налог к уплате)
        start_row = 7 + len(rows_income) + 2
        ws.cell(row=start_row, column=1, value="Раздел 1.1. Сумма налога к уплате")
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        
        rows_tax = [
            ("Код ОКТМО", "010", ""),
            ("Аванс к уплате за 1 квартал", "020", 0),
            ("Аванс к уплате за полугодие", "040", 0),
            ("Аванс к уплате за 9 месяцев", "070", 0),
            ("Налог к уплате за год", "100", self.tax_payable),
        ]
        
        for idx, (name, code, value) in enumerate(rows_tax, start_row + 2):
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=code)
            ws.cell(row=idx, column=3, value=value if value else "")
        
        # Настройка ширины
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        
        wb.save(filepath)
        
        return self.tax_payable
    
    def generate_xml(self, filepath):
        """Генерация XML-файла декларации для отправки в ЛК ФНС"""
        self.calculate()
        
        # Формируем XML по схеме ФНС
        xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Файл xmlns="urn:ФНС-СХД-Декл-УСН-2025-1" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="urn:ФНС-СХД-Декл-УСН-2025-1 ФНС-СХД-Декл-УСН-2025-1.xsd">
    <Документ>
        <КНД>1152017</КНД>
        <ДатаДок>{datetime.now().strftime('%Y-%m-%d')}</ДатаДок>
        <НомКорр>0</НомКорр>
    </Документ>
    <НалогПериод>
        <НомерПериода>34</НомерПериода>
        <ОтчетныйГод>2025</ОтчетныйГод>
    </НалогПериод>
    <Налогоплательщик>
        <ИНН>632312967829</ИНН>
        <ИП>
            <ФИО>
                <Фамилия>Леонтьев</Фамилия>
                <Имя>Артём</Имя>
                <Отчество>Владиславович</Отчество>
            </ФИО>
        </ИП>
    </Налогоплательщик>
    <Показатели>
        <Раздел1_1>
            <ОКТМО>36701320</ОКТМО>
            <СумАван010>0</СумАван010>
            <СумАван020>0</СумАван020>
            <СумАван040>0</СумАван040>
            <СумАван070>0</СумАван070>
            <СумНал100>{int(self.tax_payable)}</СумНал100>
        </Раздел1_1>
        <Раздел2_1_1>
            <СумДоход110>{int(self.quarterly_income[1])}</СумДоход110>
            <СумДоход111>{int(self.quarterly_income[1] + self.quarterly_income[2])}</СумДоход111>
            <СумДоход112>{int(self.quarterly_income[1] + self.quarterly_income[2] + self.quarterly_income[3])}</СумДоход112>
            <СумДоход113>{int(self.total_income)}</СумДоход113>
            <НалСтавка120>{self.tax_rate}</НалСтавка120>
            <СумИсчисНал130>{int(self.quarterly_income[1] * self.tax_rate / 100)}</СумИсчисНал130>
            <СумИсчисНал131>{int((self.quarterly_income[1] + self.quarterly_income[2]) * self.tax_rate / 100)}</СумИсчисНал131>
            <СумИсчисНал132>{int((self.quarterly_income[1] + self.quarterly_income[2] + self.quarterly_income[3]) * self.tax_rate / 100)}</СумИсчисНал132>
            <СумИсчисНал133>{int(self.tax_amount)}</СумИсчисНал133>
            <СумУплНал140>0</СумУплНал140>
            <СумУплНал141>0</СумУплНал141>
            <СумУплНал142>0</СумУплНал142>
            <СумУплНал143>0</СумУплНал143>
        </Раздел2_1_1>
    </Показатели>
</Файл>'''
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(xml_content)
        
        return self.tax_payable